from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import requests, json, io, os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
app = Flask(__name__) #todo the project with n8n and deploy it
CORS(app)

def get_current_week_data():
    today = datetime.date.today()
    start = today - datetime.timedelta(days=today.weekday())
    mois_fr = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    jours_fr = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    semaine_str = f"{start.day} {mois_fr[start.month - 1]} {start.year}"
    jours_list = [f"{jours_fr[(start + datetime.timedelta(days=i)).weekday()]} {(start + datetime.timedelta(days=i)).day}" for i in range(7)]
    return semaine_str, jours_list

current_semaine, current_jours = get_current_week_data()

# ── Données initiales du planning ──────────────────────────────────────
# Codes: M=Matin(8h-20h, pauses 10-11 et 17-18)  N=Nuit(20h-8h, pauses 23-00 et 4-5)
# Durée effective: 12h - 2h de pause = 10h travaillées
PLANNING_INITIAL = {
    "semaine": current_semaine,
    "jours": current_jours,
    "infirmiers": [
        {"id": 1, "nom": "Sophie Martin",    "service": "Urgences",    "gardes": ["M","M","N","R","N","N","R"]},
        {"id": 2, "nom": "Thomas Laurent",   "service": "Chirurgie",   "gardes": ["N","M","M","N","R","R","M"]},
        {"id": 3, "nom": "Léa Rousseau",     "service": "Bloc",        "gardes": ["N","R","M","M","N","M","R"]},
        {"id": 4, "nom": "Karim Benali",     "service": "Pédiatrie",   "gardes": ["R","N","N","M","M","R","N"]},
        {"id": 5, "nom": "Marie Chevalier",  "service": "Coordination","gardes": ["M","M","M","M","M","R","R"]},
        {"id": 6, "nom": "Hugo Petit",       "service": "Urgences",    "gardes": ["N","N","R","N","N","M","M"]},
        {"id": 7, "nom": "Amina Saidani",    "service": "Pédiatrie",   "gardes": ["M","R","N","N","R","N","N"]},
        {"id": 8, "nom": "Paul Durand",      "service": "Chirurgie",   "gardes": ["N","N","M","R","N","N","R"]},
    ]
}

# État en mémoire (modifiable via l'IA)
planning_state = json.loads(json.dumps(PLANNING_INITIAL))

JOURS_MAP = {"lundi":0,"lun":0,"mardi":1,"mar":1,"mercredi":2,"mer":2,
             "jeudi":3,"jeu":3,"vendredi":4,"ven":4,"samedi":5,"sam":5,"dimanche":6,"dim":6}

def planning_vers_texte():
    txt = f"Planning semaine du {planning_state['semaine']}:\n"
    for inf in planning_state["infirmiers"]:
        jours = planning_state["jours"]
        gardes = " ".join([f"{j.split()[0]}={g}" for j,g in zip(jours, inf["gardes"])])
        txt += f"- {inf['nom']} ({inf['service']}): {gardes}\n"
    return txt

# Heures effectives par garde: 12h amplitude - 2h pauses = 10h travaillées
HEURES_PAR_GARDE = {"M": 10, "N": 10, "R": 0, "C": 0}

def calculer_stats(gardes):
    heures = sum(HEURES_PAR_GARDE.get(g, 0) for g in gardes)
    nuits  = gardes.count("N")
    max_nuits_consecutives = 0
    streak = 0
    for g in gardes:
        if g == "N":
            streak += 1
            if streak > max_nuits_consecutives:
                max_nuits_consecutives = streak
        else:
            streak = 0
    alertes = []
    if heures > 60: alertes.append("⚠ Heures sup")
    if max_nuits_consecutives >= 3:  alertes.append("⚠ 3+ nuits consécutives")
    return heures, nuits, alertes

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/planning", methods=["GET"])
def get_planning():
    data = []
    for inf in planning_state["infirmiers"]:
        h, n, al = calculer_stats(inf["gardes"])
        data.append({**inf, "heures": h, "nuits": n, "alertes": al})
    return jsonify({"jours": planning_state["jours"], "semaine": planning_state["semaine"], "infirmiers": data})

@app.route("/api/planning", methods=["POST"])
def update_planning():
    body = request.json
    planning_state["infirmiers"] = body["infirmiers"]
    if "semaine" in body: planning_state["semaine"] = body["semaine"]
    return jsonify({"ok": True})

@app.route("/api/infirmier", methods=["POST"])
def add_infirmier():
    body = request.json
    nouveau = {
        "id": max([i["id"] for i in planning_state["infirmiers"]] + [0]) + 1,
        "nom": body.get("nom", "Nouveau"),
        "service": body.get("service", "Général"),
        "gardes": ["R"] * 7
    }
    planning_state["infirmiers"].append(nouveau)
    return jsonify({"ok": True})

@app.route("/api/infirmier/<int:inf_id>", methods=["PUT", "DELETE"])
def edit_delete_infirmier(inf_id):
    if request.method == "DELETE":
        planning_state["infirmiers"] = [i for i in planning_state["infirmiers"] if i["id"] != inf_id]
        return jsonify({"ok": True})
    elif request.method == "PUT":
        body = request.json
        for inf in planning_state["infirmiers"]:
            if inf["id"] == inf_id:
                if "nom" in body: inf["nom"] = body["nom"]
                if "service" in body: inf["service"] = body["service"]
                break
        return jsonify({"ok": True})

@app.route("/api/chat", methods=["POST"])
def chat():
    body      = request.json
    message   = body.get("message", "")
    api_key   = body.get("api_key", "")
    historique = body.get("historique", [])

    if not api_key:
        return jsonify({"erreur": "Clé API Groq manquante"}), 400

    planning_txt = planning_vers_texte()

    system_prompt = f"""Tu es un assistant expert en planification infirmière intégré dans une application web Excel.
Tu peux lire et modifier le planning directement.

{planning_txt}

HORAIRES (shifts de 12h avec 2h de pause, soit 10h effectives):
- M = Matin  : 8h00 - 20h00 | Pauses : 10h-11h et 17h-18h
- N = Nuit   : 20h00 - 8h00 | Pauses : 23h-00h et 4h-5h
- R = Repos
- C = Congé

CALCUL DES HEURES:
- Chaque garde (M ou N) = 10h effectives travaillées (12h amplitude - 2h pauses)
- Maximum 60h effectives/semaine par infirmier (6 gardes)

RÈGLES:
- Minimum 12h de repos entre deux gardes (shifts de 12h)
- Maximum 60h/semaine par infirmier (6 gardes max sur 7 jours)
- Pas plus de 3 nuits consécutives
- Toujours au moins 1 infirmier de nuit (code N)
- Toujours au moins 1 infirmier de jour (code M)

Quand l'utilisateur demande une modification, réponds avec:
1. Une courte explication en français
2. Un bloc JSON d'actions entre ```json et ``` avec ce format exact:
[
  {{"action": "set",  "infirmier": "Prénom Nom", "jour_index": 0, "garde": "M"}},
  {{"action": "move", "infirmier": "Prénom Nom", "de": 0, "vers": 2}},
  {{"action": "swap", "infirmier1": "Nom1", "infirmier2": "Nom2", "jour_index": 3}}
]

jour_index: 0=Lundi, 1=Mardi, 2=Mercredi, 3=Jeudi, 4=Vendredi, 5=Samedi, 6=Dimanche

Si c'est une question sans modification, réponds en français sans JSON.
Sois concis et professionnel."""

    messages = [{"role": "system", "content": system_prompt}]
    for h in historique[-6:]:  # garder les 6 derniers échanges
        messages.append(h)
    messages.append({"role": "user", "content": message})

    try:
        resp = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={"model": "llama-3.3-70b-versatile", "messages": messages,
                  "max_tokens": 800, "temperature": 0.2},
            timeout=30
        )
        if resp.status_code == 401:
            return jsonify({"erreur": "Clé API invalide. Vérifiez sur console.groq.com"}), 401
        if resp.status_code == 429:
            return jsonify({"erreur": "Limite API atteinte. Réessayez dans quelques secondes."}), 429
        resp.raise_for_status()

        contenu = resp.json()["choices"][0]["message"]["content"]

        # Extraire JSON d'actions
        actions = []
        import re
        match = re.search(r'```json\s*([\s\S]*?)\s*```', contenu)
        if match:
            try:
                actions = json.loads(match.group(1))
            except: pass

        # Texte sans JSON
        texte = re.sub(r'```json[\s\S]*?```', '', contenu).strip()

        # Appliquer les actions
        modifications = []
        for act in actions:
            result = appliquer_action(act)
            if result: modifications.append(result)

        return jsonify({
            "texte": texte,
            "actions": actions,
            "modifications": modifications,
            "planning_mis_a_jour": planning_vers_texte() if actions else None
        })

    except requests.exceptions.ConnectionError:
        return jsonify({"erreur": "Impossible de contacter l'API Groq. Vérifiez votre connexion."}), 503
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500

def appliquer_action(act):
    action = act.get("action", "")
    infirmiers = planning_state["infirmiers"]

    def trouver(nom):
        nom_lower = nom.lower()
        for inf in infirmiers:
            if nom_lower in inf["nom"].lower() or inf["nom"].lower().split()[0] in nom_lower:
                return inf
        return None

    if action == "set":
        inf = trouver(act.get("infirmier",""))
        idx = act.get("jour_index")
        garde = act.get("garde","").upper()
        if inf and idx is not None and 0 <= idx <= 6 and garde in ["M","N","R","C"]:
            ancien = inf["gardes"][idx]
            inf["gardes"][idx] = garde
            return f"{inf['nom']} : {planning_state['jours'][idx]} → {ancien} ➜ {garde}"

    elif action == "move":
        inf = trouver(act.get("infirmier",""))
        src = act.get("de"); dst = act.get("vers")
        if inf and src is not None and dst is not None:
            garde_val = inf["gardes"][src]
            inf["gardes"][dst] = garde_val
            inf["gardes"][src] = "R"
            return f"{inf['nom']} : déplacé {planning_state['jours'][src]}→{planning_state['jours'][dst]} ({garde_val})"

    elif action == "swap":
        inf1 = trouver(act.get("infirmier1",""))
        inf2 = trouver(act.get("infirmier2",""))
        idx  = act.get("jour_index")
        if inf1 and inf2 and idx is not None:
            inf1["gardes"][idx], inf2["gardes"][idx] = inf2["gardes"][idx], inf1["gardes"][idx]
            return f"Échange {inf1['nom']} ↔ {inf2['nom']} le {planning_state['jours'][idx]}"

    return None

@app.route("/api/export", methods=["POST"])
def export_excel():
    body = request.json or {}
    titre = body.get("titre", f"Planning {planning_state['semaine']}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C5"

    def fill(h): return PatternFill("solid", fgColor=h)
    def fnt(h, bold=False, sz=11): return Font(name="Arial", color=h, bold=bold, size=sz)
    def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
    def thin():
        s = Side(style="thin", color="D3D1C7")
        return Border(left=s, right=s, top=s, bottom=s)

    TEAL="0F6E56"; TEALM="1D9E75"; TEALL="E1F5EE"
    PUR="534AB7";  PURL="EEEDFE"
    AMB="854F0B";  AMBL="FAEEDA"
    BLU="185FA5";  BLUL="E6F1FB"
    NGT="2C2C2A";  NGTL="D3D1C7"
    RED="A32D2D";  REDL="FCEBEB"
    GRY="444441";  GRYM="888780"
    WHT="FFFFFF";  BG="FAFAF8"

    # Fond général
    for r in ws.iter_rows(1, 52, 1, 14):
        for c in r: c.fill = fill(BG)

    # Titre
    ws.merge_cells("A1:L1")
    ws["A1"] = f"🏥  {titre}"
    ws["A1"].font = Font(name="Arial", color=WHT, bold=True, size=13)
    ws["A1"].fill = fill(TEAL); ws["A1"].alignment = ctr()
    ws.row_dimensions[1].height = 30

    # Sous-titre horaires
    ws.merge_cells("A2:L2")
    ws["A2"] = "Shifts 12h  |  Matin : 8h00-20h00 (pauses 10h-11h, 17h-18h)  |  Nuit : 20h00-8h00 (pauses 23h-00h, 4h-5h)  |  10h effectives/shift"
    ws["A2"].font = Font(name="Arial", color=GRYM, bold=False, size=9)
    ws["A2"].fill = fill(TEALL); ws["A2"].alignment = ctr()
    ws.row_dimensions[2].height = 16

    # Légende
    legendes = [("M","Matin 8-20h",BLUL,BLU),("N","Nuit 20-8h",NGT,NGTL),
                ("R","Repos",TEALL,TEAL),("C","Congé",AMBL,AMB)]
    ws.merge_cells("A3:B3")
    ws["A3"] = "LÉGENDE :"; ws["A3"].font = fnt(GRYM, True, 9); ws["A3"].fill = fill(BG)
    col = 3
    for code, label, bg, fg in legendes:
        c = ws.cell(3, col, f"{code}={label}")
        c.fill = fill(bg); c.font = fnt(fg, True, 9); c.alignment = ctr()
        ws.column_dimensions[get_column_letter(col)].width = 11
        col += 1
    ws.row_dimensions[3].height = 16

    # Headers
    hdrs = ["#","Infirmier","Service"] + planning_state["jours"] + ["Total H","Nuits","Alertes"]
    wcol = [4, 20, 14, 9, 9, 9, 9, 9, 9, 9, 9, 7, 20]
    for i, (h, w) in enumerate(zip(hdrs, wcol), 1):
        c = ws.cell(4, i, h)
        c.fill = fill(TEALM); c.font = fnt(WHT, True, 10)
        c.alignment = ctr(); c.border = thin()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 26

    # Styles par code de garde (suppression du code S)
    sty = {"M":(BLUL,BLU),"N":(NGT,NGTL),"R":(TEALL,TEAL),"C":(AMBL,AMB)}

    for idx, inf in enumerate(planning_state["infirmiers"]):
        row = 5 + idx
        ws.row_dimensions[row].height = 32
        rb = WHT if idx % 2 == 0 else "F7F6F2"

        ws.cell(row,1,idx+1).fill = fill(rb); ws.cell(row,1).font = fnt(GRYM,sz=10); ws.cell(row,1).alignment = ctr(); ws.cell(row,1).border = thin()
        ws.cell(row,2,inf["nom"]).fill = fill(rb); ws.cell(row,2).font = fnt(GRY,True,10); ws.cell(row,2).alignment = lft(); ws.cell(row,2).border = thin()
        ws.cell(row,3,inf["service"]).fill = fill(rb); ws.cell(row,3).font = fnt(GRYM,sz=9); ws.cell(row,3).alignment = ctr(); ws.cell(row,3).border = thin()

        heures, nuits, alertes = calculer_stats(inf["gardes"])
        for d, g in enumerate(inf["gardes"]):
            bg, fg = sty.get(g, (rb, GRY))
            # Afficher le label complet pour M et N
            label = g
            c = ws.cell(row, 4+d, label)
            c.fill = fill(bg); c.font = fnt(fg, True, 11); c.alignment = ctr(); c.border = thin()

        hbg = REDL if heures > 60 else (TEALL if heures >= 40 else AMBL)
        hfg = RED  if heures > 60 else (TEAL  if heures >= 40 else AMB)
        ws.cell(row,11,heures).fill=fill(hbg); ws.cell(row,11).font=fnt(hfg,True); ws.cell(row,11).alignment=ctr(); ws.cell(row,11).border=thin()
        nbg = AMBL if nuits >= 3 else "F7F6F2"; nfg = AMB if nuits >= 3 else GRYM
        ws.cell(row,12,nuits).fill=fill(nbg); ws.cell(row,12).font=fnt(nfg,True); ws.cell(row,12).alignment=ctr(); ws.cell(row,12).border=thin()
        al_txt = " | ".join(alertes)
        abg = REDL if "sup" in al_txt else (AMBL if al_txt else rb)
        afg = RED  if "sup" in al_txt else (AMB  if al_txt else GRYM)
        ws.cell(row,13,al_txt).fill=fill(abg); ws.cell(row,13).font=fnt(afg,sz=9); ws.cell(row,13).alignment=lft(); ws.cell(row,13).border=thin()

    # Ligne totaux
    tr = 5 + len(planning_state["infirmiers"])
    ws.row_dimensions[tr].height = 22
    ws.merge_cells(f"A{tr}:C{tr}")
    ws[f"A{tr}"] = "TOTAL PAR JOUR"
    ws[f"A{tr}"].fill = fill(TEAL); ws[f"A{tr}"].font = fnt(WHT, True, 10); ws[f"A{tr}"].alignment = ctr()
    for d in range(7):
        gardes_jour = [inf["gardes"][d] for inf in planning_state["infirmiers"]]
        m = gardes_jour.count("M"); n = gardes_jour.count("N")
        c = ws.cell(tr, 4+d, f"J:{m} N:{n}")
        c.fill=fill(TEALL); c.font=fnt(TEAL,sz=9); c.alignment=ctr(); c.border=thin()

    # Ligne récap pauses
    pr = tr + 1
    ws.row_dimensions[pr].height = 18
    ws.merge_cells(f"A{pr}:C{pr}")
    ws[f"A{pr}"] = "Pauses (non comptées)"
    ws[f"A{pr}"].fill = fill(AMBL); ws[f"A{pr}"].font = fnt(AMB, True, 9); ws[f"A{pr}"].alignment = ctr()
    ws.merge_cells(f"D{pr}:G{pr}")
    ws[f"D{pr}"] = "Jour : 10h-11h  |  17h-18h"
    ws[f"D{pr}"].fill = fill(BLUL); ws[f"D{pr}"].font = fnt(BLU, False, 9); ws[f"D{pr}"].alignment = ctr()
    ws.merge_cells(f"H{pr}:K{pr}")
    ws[f"H{pr}"] = "Nuit : 23h-00h  |  4h-5h"
    ws[f"H{pr}"].fill = fill(NGTL); ws[f"H{pr}"].font = fnt(NGT, False, 9); ws[f"H{pr}"].alignment = ctr()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"planning_{planning_state['semaine'].replace(' ','_')}.xlsx"
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/reset", methods=["POST"])
def reset():
    global planning_state
    planning_state = json.loads(json.dumps(PLANNING_INITIAL))
    return jsonify({"ok": True})

if __name__ == "__main__":
    app.run(debug=True, port=5000)